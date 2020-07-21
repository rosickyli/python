#  把各个Sheet另存为单独的Excel
from openpyxl import load_workbook,Workbook


wb = load_workbook("D:/三方20200708/csv/123.xlsx")
sheetnames = wb.sheetnames
#print (sheetnames)
for k in range(0,len(sheetnames)):
    #print (sheetnames[k])
    ws = wb.worksheets[k]
    print(ws)
    # 创建新的Excel
    wb2 = Workbook()
    # 获取当前sheet
    ws2 = wb2.active
    #两个for循环遍历整个excel的单元格内容
    for i,row in enumerate(ws.iter_rows()):
        for j,cell in enumerate(row):
            # 写入新Excel
            ws2.cell(row=i+1, column=j+1, value=cell.value)
            # 设置新Sheet的名称
            ws2.title = sheetnames[k]

    wb2.save(sheetnames[k] + ".xlsx")